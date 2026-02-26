"""
PPC Hiring Test - Excel Workbook Builder
Builds 5 files that mimic Amazon Seller Central exports.
"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(BASE, "data")
OUT  = os.path.dirname(os.path.abspath(__file__))

# ── Colour palette (Amazon-ish) ──────────────────────────────────────────────
C_HEADER_BG   = "232F3E"   # Amazon dark navy
C_HEADER_FG   = "FFFFFF"
C_SUBHDR_BG   = "37475A"   # lighter navy
C_ALT_ROW     = "F7F9FB"
C_WHITE       = "FFFFFF"
C_GREEN_BG    = "D4EDDA"
C_RED_BG      = "FDDEDE"
C_YELLOW_BG   = "FFF3CD"
C_BLUE_ACCENT = "E8F0FE"
C_BORDER      = "C8D0D8"

# ACOS thresholds
ACOS_GOOD   = 0.35
ACOS_OK     = 0.60

def header_style(ws, row_num, cols):
    hdr_fill  = PatternFill("solid", fgColor=C_HEADER_BG)
    hdr_font  = Font(bold=True, color=C_HEADER_FG, size=10)
    thin      = Side(style="thin", color=C_BORDER)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill  = hdr_fill
        cell.font  = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="FF9900"),
                             left=thin, right=thin)

def subheader_style(ws, row_num, cols):
    fill = PatternFill("solid", fgColor=C_SUBHDR_BG)
    font = Font(bold=True, color=C_HEADER_FG, size=9)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

def data_row_style(ws, row_num, cols, alt=False):
    bg = C_ALT_ROW if alt else C_WHITE
    fill = PatternFill("solid", fgColor=bg)
    thin = Side(style="thin", color=C_BORDER)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = cell.fill if cell.fill.patternType else fill
        cell.border = Border(bottom=thin, left=thin, right=thin)
        cell.alignment = Alignment(vertical="center")

def acos_fill(val_str):
    """Return PatternFill based on ACOS value string like '0.274' or '27.4%'"""
    try:
        v = float(str(val_str).replace('%','').replace('$','').strip())
        if v > 1: v = v / 100  # already a percent
        if v <= ACOS_GOOD:
            return PatternFill("solid", fgColor=C_GREEN_BG)
        elif v <= ACOS_OK:
            return PatternFill("solid", fgColor=C_YELLOW_BG)
        else:
            return PatternFill("solid", fgColor=C_RED_BG)
    except:
        return None

def pct(val):
    """Format float as percentage string"""
    try:
        v = float(str(val).replace('%','').strip())
        if v > 1: return f"{v:.1f}%"
        return f"{v*100:.1f}%"
    except:
        return val or ""

def money(val):
    try:
        return f"${float(val):,.2f}"
    except:
        return val or ""

def num(val):
    try:
        v = float(val)
        return f"{v:,.0f}"
    except:
        return val or ""

def dec2(val):
    try:
        return f"{float(val):.2f}"
    except:
        return val or ""

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ── Load data ────────────────────────────────────────────────────────────────
def load_campaigns(portfolio_filter=None):
    rows = []
    with open(os.path.join(DATA, "campaign-summary.csv"), encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if portfolio_filter is None or r["Portfolio"] == portfolio_filter:
                rows.append(r)
    return rows

def load_search_terms(portfolio_filter=None, campaign_names=None):
    rows = []
    with open(os.path.join(DATA, "search-term-sample.csv"), encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if portfolio_filter and r["Portfolio"] != portfolio_filter:
                continue
            if campaign_names and r["Campaign"] not in campaign_names:
                continue
            rows.append(r)
    return rows

def load_placements(portfolio_filter=None, campaign_names=None):
    rows = []
    with open(os.path.join(DATA, "placement-summary.csv"), encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if portfolio_filter and r["Portfolio"] != portfolio_filter:
                continue
            if campaign_names and r["Campaign"] not in campaign_names:
                continue
            rows.append(r)
    return rows

def load_rank():
    rows = []
    with open(os.path.join(DATA, "rank-tracking.csv"), encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append(r)
    return rows

# ── Title banner ─────────────────────────────────────────────────────────────
def title_row(ws, title, subtitle, cols):
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    c1 = ws.cell(row=1, column=1, value=title)
    c1.fill  = PatternFill("solid", fgColor="FF9900")
    c1.font  = Font(bold=True, size=13, color="232F3E")
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.fill  = PatternFill("solid", fgColor="37475A")
    c2.font  = Font(size=9, color="FFFFFF", italic=True)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ── CAMPAIGNS tab ────────────────────────────────────────────────────────────
CAMP_COLS = [
    ("Campaign Name", 38),
    ("Status", 9),
    ("Budget ($)", 10),
    ("Targeting", 12),
    ("Bid Strategy", 22),
    ("Impressions", 12),
    ("Clicks", 8),
    ("CTR", 7),
    ("Spend ($)", 10),
    ("CPC ($)", 8),
    ("Orders", 8),
    ("Sales ($)", 10),
    ("ACOS", 8),
    ("ROAS", 7),
]

def write_campaigns_tab(ws, campaigns, portfolio_name):
    title_row(ws, f"Campaigns — {portfolio_name}",
              "Report Period: Feb 15 – Feb 21, 2026   |   Sponsored Products   |   United States",
              len(CAMP_COLS))
    ws.row_dimensions[3].height = 30
    headers = [c[0] for c in CAMP_COLS]
    widths  = [c[1] for c in CAMP_COLS]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    header_style(ws, 3, len(headers))
    set_col_widths(ws, widths)
    freeze(ws, "A4")

    font_data = Font(size=9)
    for i, r in enumerate(campaigns):
        row = i + 4
        alt = (i % 2 == 1)
        acos_raw = r.get("ACOS", "")
        vals = [
            r["Campaign Name"],
            r["Status"],
            r["Budget"],
            r["Targeting Type"],
            r["Bidding Strategy"],
            num(r["Impressions"]),
            num(r["Clicks"]),
            pct(r["CTR"]),
            money(r["Spend"]),
            money(r["CPC"]),
            num(r["Orders"]),
            money(r["Sales"]),
            pct(acos_raw),
            dec2(r["ROAS"]),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = font_data
            c.alignment = Alignment(vertical="center",
                                    horizontal="left" if col == 1 else "center")
        data_row_style(ws, row, len(vals), alt)
        # Colour ACOS cell
        if acos_raw:
            f = acos_fill(acos_raw)
            if f:
                ws.cell(row=row, column=13).fill = f
                ws.cell(row=row, column=13).font = Font(size=9, bold=True)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16
    for r in range(4, 4 + len(campaigns)):
        ws.row_dimensions[r].height = 18

# ── SEARCH TERMS tab ─────────────────────────────────────────────────────────
ST_COLS = [
    ("Campaign Name", 36),
    ("Ad Group", 28),
    ("Customer Search Term", 32),
    ("Match Type", 11),
    ("Impressions", 12),
    ("Clicks", 8),
    ("CTR", 7),
    ("CPC ($)", 8),
    ("Spend ($)", 10),
    ("Orders", 8),
    ("Sales ($)", 10),
    ("ACOS", 8),
    ("ROAS", 7),
    ("Conv. Rate", 10),
]

def write_search_terms_tab(ws, terms, portfolio_name):
    title_row(ws, f"Search Terms — {portfolio_name}",
              "Top search terms by spend   |   Report Period: Feb 15 – Feb 21, 2026",
              len(ST_COLS))
    headers = [c[0] for c in ST_COLS]
    widths  = [c[1] for c in ST_COLS]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    header_style(ws, 3, len(headers))
    set_col_widths(ws, widths)
    freeze(ws, "A4")
    font_data = Font(size=9)
    for i, r in enumerate(terms):
        row = i + 4
        alt = (i % 2 == 1)
        acos_raw = r.get("ACOS", "")
        vals = [
            r["Campaign"],
            r["Ad Group"],
            r["Search Term"],
            r.get("Match Type", ""),
            num(r["Impressions"]),
            num(r["Clicks"]),
            pct(r["CTR"]),
            money(r["CPC"]),
            money(r["Spend"]),
            num(r["Orders"]),
            money(r["Sales"]),
            pct(acos_raw),
            dec2(r["ROAS"]),
            pct(r["Conversion Rate"]),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = font_data
            c.alignment = Alignment(vertical="center",
                                    horizontal="left" if col in (1,2,3) else "center")
        data_row_style(ws, row, len(vals), alt)
        if acos_raw:
            f = acos_fill(acos_raw)
            if f:
                ws.cell(row=row, column=12).fill = f
                ws.cell(row=row, column=12).font = Font(size=9, bold=True)
    for r in range(4, 4 + len(terms)):
        ws.row_dimensions[r].height = 17

# ── PLACEMENTS tab ───────────────────────────────────────────────────────────
PL_COLS = [
    ("Campaign Name", 36),
    ("Bid Strategy", 22),
    ("Placement", 24),
    ("Impressions", 12),
    ("Clicks", 8),
    ("CPC ($)", 8),
    ("Spend ($)", 10),
    ("Sales ($)", 10),
    ("ACOS", 8),
    ("ROAS", 7),
    ("Orders", 8),
]

def write_placements_tab(ws, placements, portfolio_name):
    title_row(ws, f"Placements — {portfolio_name}",
              "Top-of-Search vs Product Pages vs Off-Amazon   |   Feb 15 – Feb 21, 2026",
              len(PL_COLS))
    headers = [c[0] for c in PL_COLS]
    widths  = [c[1] for c in PL_COLS]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    header_style(ws, 3, len(headers))
    set_col_widths(ws, widths)
    freeze(ws, "A4")
    font_data = Font(size=9)
    for i, r in enumerate(placements):
        row = i + 4
        alt = (i % 2 == 1)
        acos_raw = r.get("ACOS", "")
        vals = [
            r["Campaign"],
            r["Bidding Strategy"],
            r["Placement"],
            num(r["Impressions"]),
            num(r["Clicks"]),
            money(r["CPC"]),
            money(r["Spend"]),
            money(r["Sales"]),
            pct(acos_raw),
            dec2(r["ROAS"]),
            num(r["Orders"]),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = font_data
            c.alignment = Alignment(vertical="center",
                                    horizontal="left" if col in (1,2,3) else "center")
        data_row_style(ws, row, len(vals), alt)
        if acos_raw:
            f = acos_fill(acos_raw)
            if f:
                ws.cell(row=row, column=9).fill = f
                ws.cell(row=row, column=9).font = Font(size=9, bold=True)
    for r in range(4, 4 + len(placements)):
        ws.row_dimensions[r].height = 17

# ── RANK TRACKING tab ────────────────────────────────────────────────────────
def write_rank_tab(ws, rank_rows):
    dates = [k for k in rank_rows[0].keys() if k not in ("keyword","search_volume")]
    all_cols = ["Keyword", "Monthly Searches"] + dates
    title_row(ws, "Organic Rank Tracker — Fuse Beads Product",
              "Source: DataDive Rank Radar   |   Lower number = better rank   |   Blank = not in top 100",
              len(all_cols))
    widths = [28, 16] + [8]*len(dates)
    for col, h in enumerate(all_cols, 1):
        ws.cell(row=3, column=col, value=h)
    header_style(ws, 3, len(all_cols))
    set_col_widths(ws, widths)
    freeze(ws, "C4")

    # Rank colour scale: 1=dark green, 10=yellow, 30+=red
    green  = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red    = PatternFill("solid", fgColor="FFC7CE")
    grey   = PatternFill("solid", fgColor="F0F0F0")

    font_data  = Font(size=9)
    font_bold  = Font(size=9, bold=True)
    for i, r in enumerate(rank_rows):
        row = i + 4
        alt = (i % 2 == 1)
        ws.cell(row=row, column=1, value=r["keyword"]).font = font_bold
        sv_cell = ws.cell(row=row, column=2, value=f"{int(r['search_volume']):,}")
        sv_cell.font = font_data
        sv_cell.alignment = Alignment(horizontal="center", vertical="center")
        for d_idx, d in enumerate(dates):
            col = d_idx + 3
            val = r.get(d, "")
            c = ws.cell(row=row, column=col)
            try:
                rank = int(val)
                c.value = rank
                if rank <= 5:    c.fill = green
                elif rank <= 15: c.fill = yellow
                else:            c.fill = red
            except:
                c.value = "–"
                c.fill = grey
            c.font = Font(size=9, bold=(True if val and int(val if val else 99) <= 5 else False))
            c.alignment = Alignment(horizontal="center", vertical="center")
        # alt row for kw + sv
        bg = PatternFill("solid", fgColor=C_ALT_ROW) if alt else PatternFill("solid", fgColor=C_WHITE)
        for col in (1, 2):
            ws.cell(row=row, column=col).fill = bg
        ws.row_dimensions[row].height = 18

    # legend
    leg_row = len(rank_rows) + 5
    ws.merge_cells(start_row=leg_row, start_column=1, end_row=leg_row, end_column=6)
    c = ws.cell(row=leg_row, column=1,
                value="Colour guide:  🟢 Rank 1-5 (excellent)   🟡 Rank 6-15 (good)   🔴 Rank 16+ (at risk)   ⬜ Not in top 100")
    c.font = Font(size=9, italic=True, color="555555")


# ── BUILD SCENARIO WORKBOOK ───────────────────────────────────────────────────
def build_portfolio_workbook(portfolio_key, scenario_label, filename, include_rank=False):
    camps = load_campaigns(portfolio_key)
    camp_names = {c["Campaign Name"] for c in camps}
    terms = load_search_terms(portfolio_key, camp_names)
    placements = load_placements(portfolio_key, camp_names)

    wb = Workbook()

    # Campaigns tab
    ws_c = wb.active
    ws_c.title = "Campaigns"
    write_campaigns_tab(ws_c, camps, scenario_label)
    ws_c.sheet_view.showGridLines = False

    # Search Terms tab
    ws_s = wb.create_sheet("Search Terms")
    write_search_terms_tab(ws_s, terms, scenario_label)
    ws_s.sheet_view.showGridLines = False

    # Placements tab
    ws_p = wb.create_sheet("Placements")
    write_placements_tab(ws_p, placements, scenario_label)
    ws_p.sheet_view.showGridLines = False

    # Rank tab (beads only)
    if include_rank:
        ws_r = wb.create_sheet("Rank Tracker")
        write_rank_tab(ws_r, load_rank())
        ws_r.sheet_view.showGridLines = False

    path = os.path.join(OUT, filename)
    wb.save(path)
    print(f"  Saved: {filename}  ({len(camps)} campaigns, {len(terms)} search terms, {len(placements)} placements)")
    return path


# ── ACCOUNT OVERVIEW WORKBOOK ────────────────────────────────────────────────
def build_account_overview():
    all_camps = load_campaigns()

    # Aggregate by portfolio
    from collections import defaultdict
    portfolios = defaultdict(lambda: {"impressions":0,"clicks":0,"spend":0,
                                       "sales":0,"orders":0,"campaigns":0})
    for r in all_camps:
        p = r["Portfolio"]
        try: portfolios[p]["impressions"] += float(r["Impressions"] or 0)
        except: pass
        try: portfolios[p]["clicks"]      += float(r["Clicks"] or 0)
        except: pass
        try: portfolios[p]["spend"]       += float(r["Spend"] or 0)
        except: pass
        try: portfolios[p]["sales"]       += float(r["Sales"] or 0)
        except: pass
        try: portfolios[p]["orders"]      += float(r["Orders"] or 0)
        except: pass
        portfolios[p]["campaigns"] += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Account Overview"
    ws.sheet_view.showGridLines = False

    # Title
    N_COLS = 9
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_COLS)
    t = ws.cell(row=1, column=1, value="ArtKraft — Amazon PPC Account Overview")
    t.fill = PatternFill("solid", fgColor="FF9900")
    t.font = Font(bold=True, size=14, color="232F3E")
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    s = ws.cell(row=2, column=1, value="Report Period: Feb 15 – Feb 21, 2026   |   Sponsored Products   |   United States   |   All Portfolios")
    s.fill = PatternFill("solid", fgColor="37475A")
    s.font = Font(size=9, color="FFFFFF", italic=True)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16

    # Summary KPI row
    fin_rows = []
    with open(os.path.join(DATA, "financial-summary.csv"), encoding="utf-8") as f:
        for r in csv.DictReader(f):
            fin_rows.append(r)
    feb26 = fin_rows[0]
    feb25 = fin_rows[1]

    kpi_labels = ["Total Sales","PPC Sales","Organic Sales","Ad Spend","TACoS","PPC ACOS","ROAS","PPC CPC","Conv. Rate"]
    kpi_cur    = [feb26["Total Sales"], feb26["PPC Sales"], feb26["Organic Sales"],
                  feb26["Ad Spend"], feb26["TACoS"], feb26["PPC ACOS"],
                  feb26["ROAS"], feb26["CPC"], feb26["Conversion Rate"]]
    kpi_prev   = [feb25["Total Sales"], feb25["PPC Sales"], feb25["Organic Sales"],
                  feb25["Ad Spend"], feb25["TACoS"], feb25["PPC ACOS"],
                  feb25["ROAS"], feb25["CPC"], feb25["Conversion Rate"]]

    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 36
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 20

    for col, lbl in enumerate(kpi_labels, 1):
        ws.column_dimensions[get_column_letter(col)].width = 13
        c_lbl = ws.cell(row=4, column=col, value=lbl)
        c_lbl.fill = PatternFill("solid", fgColor=C_SUBHDR_BG)
        c_lbl.font = Font(bold=True, color="FFFFFF", size=8)
        c_lbl.alignment = Alignment(horizontal="center", vertical="center")

        c_cur = ws.cell(row=5, column=col, value=kpi_cur[col-1])
        c_cur.fill = PatternFill("solid", fgColor=C_BLUE_ACCENT)
        c_cur.font = Font(bold=True, size=12, color="232F3E")
        c_cur.alignment = Alignment(horizontal="center", vertical="center")

        c_lbl2 = ws.cell(row=6, column=col, value="Feb 2026")
        c_lbl2.font = Font(size=8, color="555555", italic=True)
        c_lbl2.alignment = Alignment(horizontal="center")

        c_prev = ws.cell(row=7, column=col, value=kpi_prev[col-1])
        c_prev.font = Font(size=9, color="888888")
        c_prev.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=6, column=1).value = "Feb 2026 ↑"
    ws.cell(row=7, column=1).value = "Feb 2025"

    # Portfolio table
    PORT_HDR = ["Portfolio", "Campaigns", "Impressions", "Clicks", "Spend ($)",
                "Sales ($)", "Orders", "ACOS", "ROAS"]
    port_widths = [34, 11, 14, 10, 12, 12, 10, 10, 9]
    ws.row_dimensions[9].height = 28

    for col, h in enumerate(PORT_HDR, 1):
        ws.cell(row=9, column=col, value=h)
        ws.column_dimensions[get_column_letter(col)].width = port_widths[col-1]
    header_style(ws, 9, len(PORT_HDR))

    font_data = Font(size=9)
    sorted_ports = sorted(portfolios.items(),
                          key=lambda x: x[1]["spend"], reverse=True)
    for i, (name, d) in enumerate(sorted_ports):
        row = i + 10
        alt = (i % 2 == 1)
        spend = d["spend"]
        sales = d["sales"]
        acos_v = (spend / sales) if sales > 0 else None
        roas_v = (sales / spend) if spend > 0 else None

        vals = [
            name,
            d["campaigns"],
            f"{d['impressions']:,.0f}",
            f"{d['clicks']:,.0f}",
            f"${spend:,.2f}",
            f"${sales:,.2f}",
            f"{d['orders']:,.0f}",
            f"{acos_v*100:.1f}%" if acos_v else "—",
            f"{roas_v:.2f}" if roas_v else "—",
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = font_data
            c.alignment = Alignment(vertical="center",
                                    horizontal="left" if col == 1 else "center")
        data_row_style(ws, row, len(PORT_HDR), alt)
        if acos_v is not None:
            f = acos_fill(acos_v)
            if f:
                ws.cell(row=row, column=8).fill = f
                ws.cell(row=row, column=8).font = Font(size=9, bold=True)
        ws.row_dimensions[row].height = 18

    # Note row
    note_row = 10 + len(sorted_ports) + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=N_COLS)
    nc = ws.cell(row=note_row, column=1,
                 value="ACOS colour guide:  🟢 ≤35% (on target)   🟡 36–60% (watch)   🔴 >60% (above target)   Note: Some portfolios include strategic launch or brand-defense campaigns where high ACOS is expected.")
    nc.font = Font(size=8, italic=True, color="555555")

    path = os.path.join(OUT, "account-overview-Feb2026.xlsx")
    wb.save(path)
    print(f"  Saved: account-overview-Feb2026.xlsx  ({len(all_camps)} total campaigns, {len(sorted_ports)} portfolios)")


# ── RUN ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Building PPC test workbooks...")

    build_account_overview()

    build_portfolio_workbook(
        "Portfolio B - Beads",
        "Fuse Beads — Portfolio B",
        "scenario-1-fuse-beads.xlsx",
        include_rank=True
    )

    build_portfolio_workbook(
        "Portfolio J - Knitting",
        "Kids Knitting — Portfolio J",
        "scenario-2-knitting.xlsx",
    )

    build_portfolio_workbook(
        "Portfolio G - Catch All",
        "Catch All Auto — Portfolio G",
        "scenario-3-catch-all.xlsx",
    )

    print("\nAll files built successfully.")
