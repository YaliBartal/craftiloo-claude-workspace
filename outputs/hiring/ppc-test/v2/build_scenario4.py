"""
Scenario 4 — 4 Flowers Adult Embroidery Portfolio
Builds the Excel workbook from hardcoded data extracted from the original reports.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.dirname(os.path.abspath(__file__))

# ── Style constants ───────────────────────────────────────────────────────────
C_HEADER_BG  = "232F3E"
C_HEADER_FG  = "FFFFFF"
C_SUBHDR_BG  = "37475A"
C_ALT_ROW    = "F7F9FB"
C_WHITE      = "FFFFFF"
C_GREEN_BG   = "D4EDDA"
C_RED_BG     = "FDDEDE"
C_YELLOW_BG  = "FFF3CD"
C_BORDER     = "C8D0D8"
C_ORANGE_BG  = "FFE0B2"   # for "up and down" bid strategy warning

ACOS_GOOD = 0.35
ACOS_OK   = 0.60

def acos_fill(val):
    try:
        v = float(str(val).replace('%','').strip())
        if v > 1: v /= 100
        if v == 0: return None
        if v <= ACOS_GOOD:  return PatternFill("solid", fgColor=C_GREEN_BG)
        if v <= ACOS_OK:    return PatternFill("solid", fgColor=C_YELLOW_BG)
        return PatternFill("solid", fgColor=C_RED_BG)
    except: return None

def pct(v):
    try:
        f = float(str(v).replace('%','').strip())
        return f"{f:.1f}%" if f > 1 else f"{f*100:.1f}%"
    except: return str(v) if v else "—"

def money(v):
    try: return f"${float(v):,.2f}"
    except: return str(v) if v else "—"

def num(v):
    try: return f"{float(v):,.0f}"
    except: return str(v) if v else "—"

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(bottom=s, left=s, right=s)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_banner(ws, title, subtitle, n_cols):
    last = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last}1")
    ws.merge_cells(f"A2:{last}2")
    c1 = ws.cell(1, 1, title)
    c1.fill = PatternFill("solid", fgColor="FF9900")
    c1.font = Font(bold=True, size=13, color="232F3E")
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2 = ws.cell(2, 1, subtitle)
    c2.fill = PatternFill("solid", fgColor=C_SUBHDR_BG)
    c2.font = Font(size=9, color=C_HEADER_FG, italic=True)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16

def write_header_row(ws, row_n, headers):
    fill = PatternFill("solid", fgColor=C_HEADER_BG)
    font = Font(bold=True, color=C_HEADER_FG, size=10)
    bot  = Side(style="medium", color="FF9900")
    s    = Side(style="thin",   color=C_BORDER)
    ws.row_dimensions[row_n].height = 28
    for col, h in enumerate(headers, 1):
        c = ws.cell(row_n, col, h)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=bot, left=s, right=s)

def write_data_row(ws, row_n, values, alt=False, acos_col=None, flag_col=None, flag_val=None):
    bg = PatternFill("solid", fgColor=C_ALT_ROW if alt else C_WHITE)
    ws.row_dimensions[row_n].height = 18
    for col, val in enumerate(values, 1):
        c = ws.cell(row_n, col, val)
        c.font = Font(size=9)
        c.border = thin_border()
        c.alignment = Alignment(vertical="center",
                                horizontal="left" if col == 1 else "center")
        c.fill = bg
    # ACOS colour
    if acos_col:
        raw = values[acos_col - 1]
        f = acos_fill(raw)
        if f:
            ws.cell(row_n, acos_col).fill = f
            ws.cell(row_n, acos_col).font = Font(size=9, bold=True)
    # Flag column (e.g. "up and down" bid strategy)
    if flag_col and flag_val and flag_val.lower() in str(values[flag_col - 1]).lower():
        ws.cell(row_n, flag_col).fill = PatternFill("solid", fgColor=C_ORANGE_BG)
        ws.cell(row_n, flag_col).font = Font(size=9, bold=True, color="8B3A00")


# ══════════════════════════════════════════════════════════════════════════════
# DATA
# ══════════════════════════════════════════════════════════════════════════════

# Campaigns tab
# Cols: Campaign Name | Status | Budget | Targeting | Bid Strategy |
#       Impressions | Clicks | CTR | Spend | CPC | Orders | Sales | ACOS | ROAS
CAMPAIGNS = [
    # name, status, budget, targeting, bid_strat, imp, clicks, ctr, spend, cpc, orders, sales, acos, roas
    ("SO | Emb. Kit | ASIN-001 | SPM | CAT | Embroidery Kits",
     "ENABLED","$35","Manual","Dynamic bids - down only",
     731, 7,"0.96%","$1.80","$0.26", 0,"$0.00","—","—"),

    ("SO | Emb. Kit | ASIN-002 | SPM | SK | cross stitching kits adult beginner",
     "ENABLED","$25","Manual","Dynamic bids - down only",
     561, 0,"0.00%","$0.00","$0.00", 0,"$0.00","—","—"),

    ("SO | Emb. Kit | ASIN-002 | SPM | PT | 4 flowers",
     "ENABLED","$5","Manual","Rule-based",
     1855,31,"1.67%","$34.20","$1.10", 2,"$49.96","68.5%","1.46"),

    ("SO | Emb. Kit | ASIN-001 | SPM | PT | Brand's ASINs",
     "ENABLED","$15","Manual","Dynamic bids - down only",
     273,  6,"2.20%","$2.76","$0.46", 0,"$0.00","—","—"),

    ("SO | Emb. Kit | SPM | Defence | Brand KW",
     "ENABLED","$30","Manual","Dynamic bids - down only",
     304,  2,"0.66%","$0.76","$0.38", 0,"$0.00","—","—"),

    ("4 flowers - shield brand self target",
     "ENABLED","$10","Manual","Dynamic bids - down only",
     599, 29,"4.84%","$27.90","$0.96", 2,"$49.96","55.8%","1.79"),

    ("4 flowers - shield brand asins",
     "ENABLED","$7","Manual","Dynamic bids - down only",
     752, 18,"2.39%","$14.04","$0.78", 0,"$0.00","—","—"),

    ("4 flowers - SPM SK embroidery starter kit beginner TOS",
     "ENABLED","$17","Manual","Dynamic bids - down only",
     483,  8,"1.66%","$9.47","$1.18", 1,"$24.98","37.9%","2.64"),

    # TRAP 1 — wrong keyword category: needlepoint ≠ embroidery
    ("4 flowers - SPM SK beginner needlepoint kits for adults TOS",
     "ENABLED","$15","Manual","Dynamic bids - down only",
     749, 31,"4.14%","$35.62","$1.15", 1,"$14.99","237.6%","0.42"),

    # TRAP 2 — "up and down" bid strategy on a 126% ACOS campaign
    ("4 flowers - SPM SK easy embroidery kits for beginners TOS",
     "ENABLED","$15","Manual","Dynamic bids - up and down",
     3029,34,"1.12%","$31.62","$0.93", 1,"$24.98","126.6%","0.79"),

    ("4 flowers - SPM MK broad ultra targeted",
     "ENABLED","$12","Manual","Dynamic bids - down only",
     13815,142,"1.03%","$84.26","$0.59", 5,"$124.90","67.5%","1.48"),

    ("4 flowers - SPM SK beginners embroidery kit for adults TOS",
     "ENABLED","$20","Manual","Dynamic bids - down only",
     3334, 60,"1.80%","$38.65","$0.64", 3,"$66.94","57.7%","1.73"),

    # TRAP 3 — "up and down" + $53.60 spend + 0 orders = money on fire
    ("4 flowers - SPM SK embroidery for beginners TOS",
     "ENABLED","$11","Manual","Dynamic bids - up and down",
     4774, 52,"1.09%","$53.60","$1.03", 0,"$0.00","—","—"),
]

# Search Terms tab
# Cols: Campaign | Search Term | Match Type | Impressions | Clicks | CTR |
#       CPC | Spend | Orders | Sales | ACOS | ROAS | Conv. Rate
SEARCH_TERMS = [
    # Best performers — worth graduating
    ("4 flowers - SPM MK broad ultra targeted","embroidery kit for adults beginners","BROAD",
     2140,28,"1.31%","$0.58","$16.24", 2,"$49.96","32.5%","3.08","7.1%"),
    ("4 flowers - SPM MK broad ultra targeted","beginner embroidery kit for adults","BROAD",
     1890,22,"1.16%","$0.61","$13.42", 1,"$24.98","53.7%","1.86","4.5%"),
    ("4 flowers - SPM SK beginners embroidery kit for adults TOS","beginners embroidery kit for adults","EXACT",
     1650,30,"1.82%","$0.65","$19.50", 2,"$44.98","43.4%","2.31","6.7%"),
    ("4 flowers - SPM MK broad ultra targeted","embroidery for beginners kit","BROAD",
     1210,18,"1.49%","$0.54","$9.72", 1,"$24.98","38.9%","2.57","5.6%"),
    ("4 flowers - SPM SK embroidery starter kit beginner TOS","embroidery starter kit for beginners","EXACT",
     483, 8,"1.66%","$1.18","$9.47", 1,"$24.98","37.9%","2.64","12.5%"),

    # TRAP 1 — needlepoint terms: wrong category, real buyers, not convertable
    ("4 flowers - SPM SK beginner needlepoint kits for adults TOS","needlepoint kits for beginners adults","EXACT",
     412,18,"4.37%","$1.22","$21.96", 0,"$0.00","—","—","0.0%"),
    ("4 flowers - SPM SK beginner needlepoint kits for adults TOS","beginner needlepoint kit","EXACT",
     337,13,"3.86%","$1.05","$13.65", 1,"$14.99","91.1%","1.10","7.7%"),

    # TRAP 2 — up and down + high ACOS terms
    ("4 flowers - SPM SK easy embroidery kits for beginners TOS","easy embroidery kit","EXACT",
     1640,18,"1.10%","$0.91","$16.38", 1,"$24.98","65.6%","1.52","5.6%"),
    ("4 flowers - SPM SK easy embroidery kits for beginners TOS","embroidery kit easy","EXACT",
     890,10,"1.12%","$0.96","$9.60", 0,"$0.00","—","—","0.0%"),
    ("4 flowers - SPM SK embroidery for beginners TOS","embroidery for beginners kit adult","EXACT",
     2210,28,"1.27%","$1.04","$29.12", 0,"$0.00","—","—","0.0%"),
    ("4 flowers - SPM SK embroidery for beginners TOS","learn embroidery kit adults","EXACT",
     1150,14,"1.22%","$0.92","$12.88", 0,"$0.00","—","—","0.0%"),

    # Wasted spend — broad pulling irrelevant traffic
    ("4 flowers - SPM MK broad ultra targeted","cross stitch kit for adults","BROAD",
     980,12,"1.22%","$0.60","$7.20", 0,"$0.00","—","—","0.0%"),
    ("4 flowers - SPM MK broad ultra targeted","needlepoint canvas","BROAD",
     750, 9,"1.20%","$0.57","$5.13", 0,"$0.00","—","—","0.0%"),
    ("4 flowers - SPM MK broad ultra targeted","sewing kit for adults","BROAD",
     620, 8,"1.29%","$0.55","$4.40", 0,"$0.00","—","—","0.0%"),
]

# Placements tab
# Cols: Campaign | Bid Strategy | Placement | Impressions | Clicks | CPC |
#       Spend | Sales | ACOS | ROAS | Orders
PLACEMENTS = [
    # Note: High TOS spend on high-ACOS campaigns = expensive mistake
    ("SO | Emb. Kit | ASIN-002 | SPM | PT | 4 flowers","Rule-based","Top of Search (on-Amazon)",
     1124,22,"$1.18","$25.96","$49.96","51.9%","1.93",2),
    ("SO | Emb. Kit | ASIN-002 | SPM | PT | 4 flowers","Rule-based","Product pages on Amazon",
     680, 8,"$0.90","$7.20","$0.00","—","—",0),
    ("SO | Emb. Kit | ASIN-002 | SPM | PT | 4 flowers","Rule-based","Off Amazon",
     51,  1,"$1.04","$1.04","$0.00","—","—",0),

    ("4 flowers - shield brand self target","Dynamic bids - down only","Top of Search (on-Amazon)",
     420,22,"$1.02","$22.44","$49.96","44.9%","2.23",2),
    ("4 flowers - shield brand self target","Dynamic bids - down only","Product pages on Amazon",
     179, 7,"$0.78","$5.46","$0.00","—","—",0),

    # TRAP: needlepoint campaign burning money at TOS
    ("4 flowers - SPM SK beginner needlepoint kits for adults TOS","Dynamic bids - down only","Top of Search (on-Amazon)",
     612,27,"$1.20","$32.40","$14.99","216.1%","0.46",1),
    ("4 flowers - SPM SK beginner needlepoint kits for adults TOS","Dynamic bids - down only","Product pages on Amazon",
     137, 4,"$0.81","$3.22","$0.00","—","—",0),

    # TRAP: up and down + TOS = maximum waste
    ("4 flowers - SPM SK easy embroidery kits for beginners TOS","Dynamic bids - up and down","Top of Search (on-Amazon)",
     2210,27,"$1.04","$28.08","$24.98","112.4%","0.89",1),
    ("4 flowers - SPM SK easy embroidery kits for beginners TOS","Dynamic bids - up and down","Product pages on Amazon",
     640, 6,"$0.59","$3.54","$0.00","—","—",0),
    ("4 flowers - SPM SK easy embroidery kits for beginners TOS","Dynamic bids - up and down","Off Amazon",
     179, 1,"$0.00","$0.00","$0.00","—","—",0),

    ("4 flowers - SPM MK broad ultra targeted","Dynamic bids - down only","Top of Search (on-Amazon)",
     7640,82,"$0.64","$52.48","$99.96","52.5%","1.91",4),
    ("4 flowers - SPM MK broad ultra targeted","Dynamic bids - down only","Product pages on Amazon",
     5440,55,"$0.51","$28.05","$24.94","112.4%","0.89",1),
    ("4 flowers - SPM MK broad ultra targeted","Dynamic bids - down only","Off Amazon",
     735, 5,"$0.75","$3.73","$0.00","—","—",0),

    ("4 flowers - SPM SK beginners embroidery kit for adults TOS","Dynamic bids - down only","Top of Search (on-Amazon)",
     2890,51,"$0.67","$34.17","$66.94","51.0%","1.96",3),
    ("4 flowers - SPM SK beginners embroidery kit for adults TOS","Dynamic bids - down only","Product pages on Amazon",
     444, 9,"$0.50","$4.48","$0.00","—","—",0),

    # TRAP: $53.60 spent, 0 orders, all on TOS with up and down
    ("4 flowers - SPM SK embroidery for beginners TOS","Dynamic bids - up and down","Top of Search (on-Amazon)",
     3810,43,"$1.08","$46.44","$0.00","—","—",0),
    ("4 flowers - SPM SK embroidery for beginners TOS","Dynamic bids - up and down","Product pages on Amazon",
     752, 8,"$0.90","$7.20","$0.00","—","—",0),
    ("4 flowers - SPM SK embroidery for beginners TOS","Dynamic bids - up and down","Off Amazon",
     212, 1,"$0.00","$0.00","$0.00","—","—",0),

    ("4 flowers - SPM SK embroidery starter kit beginner TOS","Dynamic bids - down only","Top of Search (on-Amazon)",
     421, 7,"$1.21","$8.47","$24.98","33.9%","2.95",1),
    ("4 flowers - SPM SK embroidery starter kit beginner TOS","Dynamic bids - down only","Product pages on Amazon",
     62,  1,"$1.00","$1.00","$0.00","—","—",0),
]


# ══════════════════════════════════════════════════════════════════════════════
# BUILD
# ══════════════════════════════════════════════════════════════════════════════

wb = Workbook()

# ── Tab 1: Campaigns ──────────────────────────────────────────────────────────
ws_c = wb.active
ws_c.title = "Campaigns"
ws_c.sheet_view.showGridLines = False

CAMP_HDR = ["Campaign Name","Status","Budget","Targeting","Bid Strategy",
            "Impressions","Clicks","CTR","Spend ($)","CPC ($)",
            "Orders","Sales ($)","ACOS","ROAS"]
CAMP_W   = [46, 9, 9, 10, 26, 12, 8, 7, 10, 9, 8, 10, 8, 7]

title_banner(ws_c, "Campaigns — 4 Flowers Adult Embroidery  (Portfolio H)",
             "Report Period: Feb 15 – Feb 21, 2026   |   Sponsored Products   |   United States",
             len(CAMP_HDR))
write_header_row(ws_c, 3, CAMP_HDR)
set_col_widths(ws_c, CAMP_W)
ws_c.freeze_panes = "A4"

for i, row in enumerate(CAMPAIGNS):
    r = i + 4
    write_data_row(ws_c, r, list(row), alt=(i%2==1),
                   acos_col=13,   # ACOS
                   flag_col=5,    # Bid Strategy
                   flag_val="up and down")

# Summary row
sum_row = len(CAMPAIGNS) + 5
ws_c.merge_cells(f"A{sum_row}:{get_column_letter(len(CAMP_HDR))}{sum_row}")
sc = ws_c.cell(sum_row, 1,
    "Portfolio total:  Spend $304.68  |  Sales $356.71  |  Portfolio ACOS ~85%  |  13 campaigns  |  15 total orders")
sc.fill = PatternFill("solid", fgColor="EAF4FB")
sc.font = Font(size=9, bold=True, color="1A5276")
sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_c.row_dimensions[sum_row].height = 20

# ── Tab 2: Search Terms ───────────────────────────────────────────────────────
ws_s = wb.create_sheet("Search Terms")
ws_s.sheet_view.showGridLines = False

ST_HDR = ["Campaign Name","Customer Search Term","Match Type","Impressions",
          "Clicks","CTR","CPC ($)","Spend ($)","Orders","Sales ($)",
          "ACOS","ROAS","Conv. Rate"]
ST_W   = [44, 36, 12, 12, 8, 7, 9, 10, 8, 10, 8, 7, 11]

title_banner(ws_s, "Search Terms — 4 Flowers Adult Embroidery  (Portfolio H)",
             "Top search terms by spend   |   Feb 15 – Feb 21, 2026",
             len(ST_HDR))
write_header_row(ws_s, 3, ST_HDR)
set_col_widths(ws_s, ST_W)
ws_s.freeze_panes = "A4"

for i, row in enumerate(SEARCH_TERMS):
    r = i + 4
    write_data_row(ws_s, r, list(row), alt=(i%2==1), acos_col=11)

# ── Tab 3: Placements ─────────────────────────────────────────────────────────
ws_p = wb.create_sheet("Placements")
ws_p.sheet_view.showGridLines = False

PL_HDR = ["Campaign Name","Bid Strategy","Placement","Impressions",
          "Clicks","CPC ($)","Spend ($)","Sales ($)","ACOS","ROAS","Orders"]
PL_W   = [44, 26, 28, 12, 8, 9, 10, 10, 8, 7, 8]

title_banner(ws_p, "Placements — 4 Flowers Adult Embroidery  (Portfolio H)",
             "Top-of-Search vs Product Pages vs Off-Amazon   |   Feb 15 – Feb 21, 2026",
             len(PL_HDR))
write_header_row(ws_p, 3, PL_HDR)
set_col_widths(ws_p, PL_W)
ws_p.freeze_panes = "A4"

for i, row in enumerate(PLACEMENTS):
    r = i + 4
    write_data_row(ws_p, r, list(row), alt=(i%2==1),
                   acos_col=9,
                   flag_col=2, flag_val="up and down")

# ── Notes tab ─────────────────────────────────────────────────────────────────
ws_n = wb.create_sheet("Context")
ws_n.sheet_view.showGridLines = False
title_banner(ws_n, "4 Flowers — Product Context", "", 2)

notes = [
    ("Product", "4 Flowers Embroidery Kit for Adults — beginner-friendly, floral design, includes hoop, threads, needle, instructions"),
    ("Price point", "$24.99"),
    ("Launch date", "Approx. October 2025 (~4 months ago)"),
    ("Category", "Embroidery Kits > For Adults > Beginner"),
    ("Target audience", "Adult beginners, gift buyers, craft hobbyists"),
    ("Primary keywords", "embroidery kit for beginners, beginner embroidery kit for adults, embroidery starter kit"),
    ("Average order value", "$24.99 (single unit)"),
    ("Weekly sales target", "$250–$300"),
    ("Current weekly sales", "~$357 (just above target, but spending $305 to get there)"),
    ("Star rating", "4.3 stars, 28 reviews"),
    ("Main competitor ASINs", "ASIN-010, ASIN-011, ASIN-012 (avg 4.5 stars, 200–800 reviews)"),
    ("", ""),
    ("Note from manager", "This portfolio has been running for 3 months and we can't get ACOS below 80%. "
                          "We're spending almost as much as we're making. Something is structurally wrong."),
]

ws_n.column_dimensions["A"].width = 22
ws_n.column_dimensions["B"].width = 80

for i, (label, value) in enumerate(notes):
    r = i + 4
    ws_n.row_dimensions[r].height = 18
    cl = ws_n.cell(r, 1, label)
    cv = ws_n.cell(r, 2, value)
    if label:
        cl.font = Font(bold=True, size=9, color="232F3E")
        cl.fill = PatternFill("solid", fgColor="EAF4FB" if i%2==0 else C_WHITE)
        cv.fill = PatternFill("solid", fgColor="EAF4FB" if i%2==0 else C_WHITE)
    cv.font = Font(size=9)
    cv.alignment = Alignment(vertical="center", wrap_text=True)
    if label == "Note from manager":
        cv.font = Font(size=9, italic=True, color="8B3A00")
        cv.fill = PatternFill("solid", fgColor=C_ORANGE_BG)
        cl.fill = PatternFill("solid", fgColor=C_ORANGE_BG)

out = os.path.join(OUT, "scenario-4-four-flowers.xlsx")
wb.save(out)
print(f"Saved: scenario-4-four-flowers.xlsx")
print(f"  {len(CAMPAIGNS)} campaigns | {len(SEARCH_TERMS)} search terms | {len(PLACEMENTS)} placement rows")
print(f"  + Context tab with product notes")
