import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook('outputs/research/negative-keywords/data/fuse-beads-st-temp.xlsx')
ws = wb.active
headers = [cell.value for cell in ws[1]]
h = {str(v).strip(): i for i, v in enumerate(headers) if v}

portfolio_col = h['Portfolio name']
campaign_col = h['Campaign Name']
search_term_col = h['Customer Search Term']
spend_col = h['Spend']
orders_col = h['7 Day Total Orders (#)']
clicks_col = h['Clicks']
sales_col = h['7 Day Total Sales']
impressions_col = h['Impressions']

fuse_mini = []
biggie = []

for row in ws.iter_rows(min_row=2, values_only=True):
    portfolio = str(row[portfolio_col] or '').strip()
    term = str(row[search_term_col] or '').strip()
    spend = float(row[spend_col] or 0)
    orders = int(float(row[orders_col] or 0))
    clicks = int(float(row[clicks_col] or 0))
    sales = float(row[sales_col] or 0)
    impressions = int(float(row[impressions_col] or 0))
    campaign = str(row[campaign_col] or '').strip()

    entry = {'term': term, 'spend': spend, 'orders': orders, 'clicks': clicks, 'sales': sales, 'impressions': impressions, 'campaign': campaign}

    if portfolio == 'fuse beads':
        fuse_mini.append(entry)
    elif portfolio == 'biggie beads':
        biggie.append(entry)

def agg_terms(rows):
    agg = defaultdict(lambda: {'spend': 0, 'orders': 0, 'clicks': 0, 'sales': 0, 'impressions': 0, 'campaigns': set()})
    for r in rows:
        t = r['term']
        agg[t]['spend'] += r['spend']
        agg[t]['orders'] += r['orders']
        agg[t]['clicks'] += r['clicks']
        agg[t]['sales'] += r['sales']
        agg[t]['impressions'] += r['impressions']
        agg[t]['campaigns'].add(r['campaign'])
    result = []
    for term, d in agg.items():
        acos = (d['spend']/d['sales']*100) if d['sales'] > 0 else None
        result.append({'term': term, 'spend': round(d['spend'],2), 'orders': d['orders'], 'clicks': d['clicks'], 'sales': round(d['sales'],2), 'impressions': d['impressions'], 'acos': round(acos,1) if acos else None})
    return sorted(result, key=lambda x: x['spend'], reverse=True)

mini_terms = agg_terms(fuse_mini)
biggie_terms = agg_terms(biggie)

print('=== FUSE BEADS (mini 2.6mm) ===')
print(f'Rows: {len(fuse_mini)}, Unique terms: {len(mini_terms)}')
mini_spend = sum(t['spend'] for t in mini_terms)
mini_orders = sum(t['orders'] for t in mini_terms)
print(f'Spend: ${mini_spend:.2f}, Orders: {mini_orders}')
zero = [t for t in mini_terms if t['orders']==0]
zero_spend = sum(t['spend'] for t in zero)
print(f'Zero-order: {len(zero)} terms, ${zero_spend:.2f}')

print('\nP1 ($5+, 0 orders):')
for t in [x for x in mini_terms if x['orders']==0 and x['spend']>=5]:
    print(f'  ${t["spend"]} | {t["clicks"]}cl | {t["term"]}')

print('\nP2 ($2-5, 0 orders):')
for t in [x for x in mini_terms if x['orders']==0 and 2 <= x['spend'] < 5]:
    print(f'  ${t["spend"]} | {t["clicks"]}cl | {t["term"]}')

print('\nIrrelevant-looking terms (any spend):')
irrelevant_patterns = ['sewing','embroidery','cross stitch','latch','crochet','knitting','quilt','diamond','paint by','pottery','clay','slime','tie dye','scrapbook','origami','jewelry','necklace','bracelet','pillow','wallet','pencil case','needle','hook','stitch','fabric','thread','yarn','felt']
for t in mini_terms:
    term_lower = t['term'].lower()
    for pat in irrelevant_patterns:
        if pat in term_lower:
            print(f'  ${t["spend"]} | {t["clicks"]}cl | {t["orders"]}ord | {t["term"]} [matches: {pat}]')
            break

print('\n=== BIGGIE BEADS (10mm) ===')
print(f'Rows: {len(biggie)}, Unique terms: {len(biggie_terms)}')
big_spend = sum(t['spend'] for t in biggie_terms)
big_orders = sum(t['orders'] for t in biggie_terms)
print(f'Spend: ${big_spend:.2f}, Orders: {big_orders}')
zero_b = [t for t in biggie_terms if t['orders']==0]
zero_b_spend = sum(t['spend'] for t in zero_b)
print(f'Zero-order: {len(zero_b)} terms, ${zero_b_spend:.2f}')

print('\nP1 ($5+, 0 orders):')
for t in [x for x in biggie_terms if x['orders']==0 and x['spend']>=5]:
    print(f'  ${t["spend"]} | {t["clicks"]}cl | {t["term"]}')

print('\nAll biggie terms with $1+ spend:')
for t in [x for x in biggie_terms if x['spend'] >= 1]:
    print(f'  ${t["spend"]} | {t["clicks"]}cl | {t["orders"]}ord | ACoS:{t["acos"]}% | {t["term"]}')

print('\nIrrelevant-looking terms in biggie (any spend):')
for t in biggie_terms:
    term_lower = t['term'].lower()
    for pat in irrelevant_patterns:
        if pat in term_lower:
            print(f'  ${t["spend"]} | {t["clicks"]}cl | {t["orders"]}ord | {t["term"]} [matches: {pat}]')
            break

print('\n=== CROSS-PORTFOLIO ===')
print('Mini portfolio terms mentioning biggie/large/10mm:')
for t in mini_terms:
    tl = t['term'].lower()
    if any(w in tl for w in ['10mm','biggie','big bead','large bead','jumbo','big fuse','large fuse','large perler']):
        print(f'  ${t["spend"]} | {t["clicks"]}cl | {t["orders"]}ord | {t["term"]}')

print('\nBiggie portfolio terms mentioning mini/2.6mm:')
for t in biggie_terms:
    tl = t['term'].lower()
    if any(w in tl for w in ['mini','2.6','small bead','tiny','2.6mm']):
        print(f'  ${t["spend"]} | {t["clicks"]}cl | {t["orders"]}ord | {t["term"]}')

print('\n=== TOP 30 FUSE BEADS TERMS BY SPEND ===')
for i, t in enumerate(mini_terms[:30]):
    acos_str = f'{t["acos"]}%' if t["acos"] else 'N/A'
    print(f'{i+1}. ${t["spend"]} | {t["clicks"]}cl | {t["orders"]}ord | ACoS:{acos_str} | {t["term"]}')

wb.close()
